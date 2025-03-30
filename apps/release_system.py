from fastapi import APIRouter,Request,HTTPException
import db.mysql as Mysql
from models.config import *
import json
from classes.landun import Landun
from classes.gitwoa import GitWoa
from classes.excel_compare import ExcelCompare
from classes.compatible import Compatible
from datetime import datetime
import logging
import asyncio
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import load_workbook
from classes.config_api import ConfigApi

Router = APIRouter(    
    prefix="/api/v1",
    tags=["release_system"],
    responses={404: {"description": "Not found"}},)

@Router.post("/baseCICheck")
async def baseCICheckResponese(param: BaseCIParams):
    event = param.event

    config,errorInfo = await Mysql.get_datas(Config,{'fieldKey':'base_app_auto_merge'},True)
    
    branch_list = json.loads(config.strContent)
    
    branch = event.ref.split('/')[-1]
    
    match = False
    
    if branch in branch_list:
        match = True
        
    return {'match': match}

@Router.post("/modifyPipelineParamValues")
async def modifyPipelineParamValuesResponese(param: ModifyPipelineParams):
    
    pipelineUrl = param.pipeline_url
    params = param.params
    type = param.type
    config,errorInfo = await Mysql.get_datas(Config,{'fieldKey':'landun_token'},True)
    
    landun = Landun(config.strContent)
    
    res = await landun.modify_pipeline([pipelineUrl],params,type)
    if res:
        return {'status': 0,'msg': "success"}
    else:
        return {'status': -1,'msg': "modifyPipelineParamValues failed"}

@Router.options("/getPipelineParams")
@Router.post("/getPipelineParams")
async def getPipelineParamsResponese(param: GetPipelineParams,request: Request):
    if request.method == 'OPTIONS':
        return {'msg': "success"}
    
    pipelineUrl = param.pipeline_url
    params = param.params
    config,errorInfo = await Mysql.get_datas(Config,{'fieldKey':'landun_token'},True)
    
    landun = Landun(config.strContent)
    
    res = await landun.get_pipeline_params(pipelineUrl,params)
    
    return {'msg': "success","data": res}

@Router.options("/getPipelineParamsByHistory")
@Router.post("/getPipelineParamsByHistory")
async def getPipelineParamsByHistoryResponese(param: GetPipelineParamsByHistory,request: Request):
    if request.method == 'OPTIONS':
        return {'msg': "success"}
    
    pipelineUrl = param.pipeline_url
    params = param.params
    config,errorInfo = await Mysql.get_datas(Config,{'fieldKey':'landun_token'},True)
    
    landun = Landun(config.strContent)
    
    res = await landun.get_pipeline_params_by_history(pipelineUrl,params)
    
    return {'msg': "success","data": res}

@Router.post("/setExcelCompareInfo")
async def setExcelCompareInfo(param: ExcelCompareParams):
    

    ctime = datetime.now()
    compare = YmzxExcelCompare()
    compare.excel_name = param.excel_name
    compare.self_branch = param.self_branch
    compare.compare_branch = param.compare_branch
    compare.sheet_names = json.dumps(param.sheet_names,ensure_ascii=False)
    compare.compare_sheet_names = json.dumps(param.compare_sheet_names,ensure_ascii=False)
    compare.diff_infos = json.dumps(param.diff_infos,ensure_ascii=False)
    compare.self_commit = param.self_commit
    compare.date = datetime.strptime(param.date,"%a %b %d %H:%M:%S %Y %z")
    compare.author = param.author
    compare.msg = param.msg
    compare.compare_commit = param.compare_commit
    compare.compare_date = datetime.strptime(param.compare_date,"%a %b %d %H:%M:%S %Y %z")
    compare.compare_author = param.compare_author
    compare.compare_msg = param.compare_msg
    compare.full_path = param.full_path
    compare.ctime = ctime
    
    
    compare.ctime = ctime
    #logging.getLogger('INFO').info(f'YmzxExcelCompare: {compare}')
    
    data,errorInfo = await Mysql.get_datas(YmzxExcelCompare,{'excel_name':compare.excel_name,'self_branch':compare.self_branch,'compare_branch':compare.compare_branch},True)
    
    if data:
        res,errorInfo = await Mysql.update_data(YmzxExcelCompare,{'excel_name':compare.excel_name,'self_branch':compare.self_branch,'compare_branch':compare.compare_branch},compare.model_dump(exclude_unset=True))
    else:
        res,errorInfo = await Mysql.insert(YmzxExcelCompare,[compare])
        
    if not res:
        return {'msg': "failed","error": errorInfo}
    
    return {'msg': "success"}

@Router.post("/getExcelCompareInfo")
async def getExcelCompareInfo(param: ExcelCompareParams):
    
    data,errorInfo = await Mysql.get_datas(YmzxExcelCompare,{'excel_name':param.excel_name,'self_branch':param.self_branch,'compare_branch':param.compare_branch},True)
    
    if data:
        return {'msg': "success","data": data.model_dump()}
    else:
        return {'msg': "failed","error": errorInfo}
    
@Router.post("/getExcelOriginData")
async def getExcelOriginData(param: ExcelOriginParams):

    woa = GitWoa()
    result = await woa.getFileContent('TimiT1/MOE/LetsGo/letsgo_common',param.full_path,param.commit)
    if result:
        return StreamingResponse(BytesIO(result), media_type="application/octet-stream")
    else:
        raise HTTPException(status_code=404, detail="Item not found")
    
@Router.post("/getExcelData")
async def getExcelOriginData(param: ExcelOriginParams):

    woa = GitWoa()
    result = await woa.getFileContent('TimiT1/MOE/LetsGo/letsgo_common',param.full_path,param.commit)
    if result:
        file_stream = BytesIO(result)
        excel = load_workbook(file_stream,data_only=True)
        res = []
        for sheet in excel.sheetnames:
            data_array = []
            for row in excel[sheet].iter_rows(values_only=True):
                data_array.append(list(row))
            res.append({'sheet': sheet,'data': data_array})
        return {'msg': "success","data": res}
    else:
        raise HTTPException(status_code=404, detail="Item not found")
    
    
@Router.post("/addExcelCompareRecord")
async def addExcelCompareRecord(param: ExcelCompareParamsRecord):
    
    ctime = datetime.now()
    compare = YmzxExcelCompareRecord()
    compare.excels = json.dumps(param.excels,ensure_ascii=False)
    compare.self_branch = param.self_branch
    compare.compare_branch = param.compare_branch
    compare.ctime = ctime
    rid = int(param.rid)
    compare.is_finish = 1
    
    #logging.getLogger('INFO').info(f'ExcelCompareParamsRecord: {compare}')
    
    res,errorInfo = await Mysql.update_data(YmzxExcelCompareRecord,{'rid':rid},compare.model_dump(exclude_unset=True))

    if not res:
        return {'msg': "failed","error": errorInfo}
    
    return {'msg': "success"}

@Router.post("/getExcelCompareRecord")
async def getExcelCompareRecord(param: ExcelCompareParamsRecordQuery):
    data,errorInfo = await Mysql.get_datas(YmzxExcelCompareRecord,{'rid':param.rid},True)
    if data:
        excelLists,errorInfo = await Mysql.get_datas(YmzxExcelCompare,{'self_branch':data.self_branch,'compare_branch':data.compare_branch})
        excels_res = {}
        for excel in excelLists:
            diffs = json.loads(excel.diff_infos)
            if diffs['addsheet'] == False and not diffs['sheetsInfo']:
                excels_res[excel.full_path] = 'NoChange'
            elif diffs['addsheet'] == True:
                excels_res[excel.full_path] = 'New'
            elif diffs['addsheet'] == False and diffs['sheetsInfo']:
                excels_res[excel.full_path] = 'Change'
        return {'msg': "success","data": excels_res}
    else:
        return {'msg': "failed","error": errorInfo}
    
    
    
@Router.post("/getCurCompatibleInfo")
async def getCurCompatibleInfo(param: CurCompatibleParams):
    
    woa = GitWoa()
    result = await woa.getFileContent('TimiT1/MOE/LetsGo/letsgo_common',"excel/xls/B_版本兼容组.xlsx",param.branch)
    if result:
        file_stream = BytesIO(result)
        excel = load_workbook(file_stream,data_only=True)
        dsMap = {}
        for sheet in excel.sheetnames:
            compatible = Compatible(sheet,excel)
            maxVersion = compatible.get_max_client_ver()
            dsMap = dsMap | compatible.get_ds_info(maxVersion)
        #logging.getLogger("INFO").info(f'dsMap: {dsMap}')
        tasks = [Mysql.get_datas(Config,{'fieldKey':'release_play_module','tenant_id':'master'},True),Mysql.get_datas(Config,{'fieldKey':'release_play_group','tenant_id':'master'},True)]
        results = await asyncio.gather(*tasks)
        if results[0][0] and results[1][0]:
            release_play_module = json.loads(results[0][0].strContent)
            release_play_group = json.loads(results[1][0].strContent)
            module_list = {}
            for value in release_play_module:
                if value['play_ids'] == "":
                    module_list[value['play_module_name']] = []
                else:
                    module_list[value['play_module_name']] = value['play_ids'].split(',')

            module_map = {}
            for value in release_play_group:
                for module_name in value['play_modules']:
                    module_map[module_name] = value['play_group_name']

            res = {}
            
            for moduleName,playIds in module_list.items():
                if not playIds:
                    res[moduleName] = dsMap[module_map[moduleName]]
                else:
                    playId = playIds[0]
                    if playId in dsMap:
                        res[moduleName] = dsMap[playId]
                    else:
                        res[moduleName] = None
                    
            return {'msg': "success","data": res}
        else:
            return {'msg': "failed","error": "mysql callback error"}

    else:
        raise HTTPException(status_code=404, detail="Item not found")
    
@Router.post("/doExcelCompare")
async def doExcelCompare(param: ExcelCompareApiParam):
    
    woa = GitWoa()
    file1 = await woa.getFileContent('TimiT1/MOE/LetsGo/letsgo_common',param.full_path,param.self_branch)
    file2 = await woa.getFileContent('TimiT1/MOE/LetsGo/letsgo_common',param.full_path,param.compare_branch)
    if file1 and file2:
        file_stream1 = BytesIO(file1)
        file_stream2 = BytesIO(file2)
        excel1 = load_workbook(file_stream1,data_only=True)
        excel2 = load_workbook(file_stream2,data_only=True)
        excel_compare = ExcelCompare(excel1,excel2)
        res = excel_compare.compare_excel()
    else:
        raise HTTPException(status_code=404, detail="Item not found")
    
@Router.post("/getConfigSettings")
async def getConfigSettings(param:GetConfigSettings):
    
    res,type = await ConfigApi().getSettings(param.fieldKey)
    
    return {'msg': "success","data": res,"type":type}


@Router.post("/setWeddingInfo")
async def setWeddingInfo(param:SetWeddingInfo):
    
    res,errorInfo = await Mysql.get_datas(WeddingInfo,{'name':param.name},True)
    if res:
        rres,errorInfo = await Mysql.update_data(WeddingInfo,{'name':param.name},param.__dict__)
    else:
        rres,errorInfo = await Mysql.insert(WeddingInfo,[WeddingInfo(**param.__dict__)])
    return {'msg': "success","data": rres}

@Router.post("/getConfigSettings")
async def getConfigSettings(param:GetConfigSettings):
    
    res,errorInfo = await Mysql.get_datas(WeddingInfo,{'name':param.name},True)
    if res:
        return {'msg': "success","data": res}
    return {'msg': "error","data": errorInfo}