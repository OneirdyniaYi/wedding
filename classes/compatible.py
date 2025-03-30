import openpyxl
class Compatible:
    def __init__(self,sheet,excel:openpyxl.Workbook):
        self.table = excel[sheet]
        self.sheetName = sheet
        self.excel = excel
    
    def get_client_ver_data(self, client_ver):
        rows_data=[]
        for i in range(4,self.table.max_row+1):
            client_ver_tmp=self.table.cell(i,1).value
            if client_ver_tmp is None or client_ver_tmp != client_ver:
                    continue
            one_row = []
            for j in range(1, self.table.max_column+1):
                one_row.append(self.table.cell(i,j).value)
            rows_data.append(one_row)
        return rows_data
    
    def get_max_client_ver(self):
        max_client_ver='0.0.0.0'
        for i in range(4,self.table.max_row+1):
            client_ver_tmp=self.table.cell(i,1).value
            if client_ver_tmp is None:
                continue

            if list(map(int, max_client_ver.split('.'))) < list(map(int, client_ver_tmp.split('.'))):
                max_client_ver=client_ver_tmp
        return max_client_ver
    
    def get_ds_info(self,version):
        res = {}
        rows = self.get_client_ver_data(version)
        if not rows:
            return
        for row in rows:
            mt = row[3].split(',') if row[3] else []
            ds_version_record = row[2]
            if self.sheetName == "版本号控制":
                ds_version_record = row[1]
            if not mt:
                res[self.sheetName] = {"sheet":self.sheetName,"client_version":row[0],"ds_version":ds_version_record}
            else:
                for playId in mt:
                    res[playId] = {"sheet":self.sheetName,"client_version":row[0],"ds_version":ds_version_record}
        return res