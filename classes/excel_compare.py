import openpyxl
import os
import difflib

class ExcelCompare:
    def __init__(self,file1:openpyxl.Workbook,file2:openpyxl.Workbook):
        self.file1 = file1
        self.file2 = file2
        self.common_sheets = set(self.file1.sheetnames) & set(self.file2.sheetnames)
        self.diff_sheets = set(self.file1.sheetnames) ^ set(self.file2.sheetnames)
        self.add_sheets = set(self.file1.sheetnames) - set(self.file2.sheetnames)
        self.delete_sheets = set(self.file2.sheetnames) - set(self.file1.sheetnames)
    
    def read_excel_sheet(self,wb, sheet_name):
        # 获取工作表
        ws = wb[sheet_name]

        # 将工作表转换为列表，每一行是一个列表，其中每个元素是一个元组，包含单元格的坐标和值
        lines = [[(cell.coordinate, cell.value) if cell.value is not None else (cell.coordinate, "") for cell in row] for row in ws]

        return lines
    
    def _fancy_replace_non_recursive(self, a, alo, ahi, b, blo, bhi, modified_lines, added_lines, deleted_lines, lines1, lines2):
        cruncher = difflib.SequenceMatcher(lambda s: '\t' in s)
        
        # 使用栈来模拟递归
        stack = [(alo, ahi, blo, bhi)]
        
        while stack:
            alo, ahi, blo, bhi = stack.pop()
            if alo < ahi and blo < bhi:
                best_ratio, cutoff = 0.74, 0.75
                eqi, eqj = None, None   # 1st indices of equal lines (if any)
                best_i, best_j = None, None 
                

                for j in range(blo, bhi):
                    if best_ratio >= 0.85:
                        break
                    bj = b[j]
                    cruncher.set_seq2(bj)
                    for i in range(alo, ahi):
                        ai = a[i]
                        if ai == bj:
                            if eqi is None:
                                eqi, eqj = i, j
                                best_ratio, best_i, best_j = 1.0, i, j
                                break
                        cruncher.set_seq1(ai)
                        if cruncher.real_quick_ratio() > best_ratio and \
                            cruncher.quick_ratio() > best_ratio and \
                            cruncher.ratio() > best_ratio:
                            best_ratio, best_i, best_j = cruncher.ratio(), i, j
                            if best_ratio >= 0.85:
                                break

                if best_ratio < cutoff:
                    if eqi is None:
                        deleted_lines.update({i+1: lines1[i] for i in range(alo, ahi)})
                        added_lines.update({j+1: lines2[j] for j in range(blo, bhi)})
                        continue
                    best_i, best_j, best_ratio = eqi, eqj, 1.0
                else:
                    eqi = None
                # 模拟递归调用的前半部分
                if alo < best_i or blo < best_j:
                    stack.append((alo, best_i, blo, best_j))

                if eqi is None:
                    old_line = lines1[best_i]
                    new_line = lines2[best_j]
                    if old_line != new_line:
                        modified_cells = []
                        max_len = max(len(old_line), len(new_line))
                        for i in range(max_len):
                            if i < len(old_line)and i < len(new_line) and old_line[i][1] != new_line[i][1]:
                                modified_cells.append((old_line[i][0],old_line[i][1], new_line[i][1]))
                            elif i < len(old_line) and i >= len(new_line) and old_line[i][1] != '':
                                modified_cells.append((old_line[i][0],old_line[i][1], ''))
                            elif i >= len(old_line) and i < len(new_line) and new_line[i][1] != '':
                                obj = re.sub(r'\d+', '', new_line[i][0]) + str(i+1)
                                modified_cells.append((obj,'', new_line[i]))
                        if modified_cells:
                            modified_lines[best_i+1] = modified_cells

                # 模拟递归调用的后半部分
                if best_i + 1 < ahi or best_j + 1 < bhi:
                    stack.append((best_i + 1, ahi, best_j + 1, bhi))
            elif alo < ahi:
                # blo >= bhi, 处理删除的情况
                deleted_lines.update({i+1: lines1[i] for i in range(alo, ahi)})
            elif blo < bhi:
                # alo >= ahi, 处理添加的情况
                added_lines.update({j+1: lines2[j] for j in range(blo, bhi)})
    
    
    def diff_excel_sheet(self, sheet_name):
        # 读取两个Excel文件中的同一个表格
        lines1 = self.read_excel_sheet(self.file1, sheet_name)
        lines2 = self.read_excel_sheet(self.file2, sheet_name)

        seq1 = ['\t'.join([str(cell[1]) for cell in line]) for line in lines1]
        seq2 = ['\t'.join([str(cell[1]) for cell in line]) for line in lines2]
        # 使用Differ对象比较两个表格
        matcher = difflib.SequenceMatcher(None,seq1,seq2)

        # print("------------------------------------seq1------------------------------------")
        # print(seq1)
        # print("------------------------------------seq2------------------------------------")
        # print(seq2)
        # print("------------------------------------done--------------------------------------")
        
        
        print(f"----------------------compare sheet: {sheet_name}----------------------")
        # 初始化修改的行的字典
        modified_lines = {}
        added_lines = {}
        deleted_lines = {}

        # 遍历差异，查找修改的行
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            print(tag, i1, i2, j1, j2)
            if tag == 'replace':
                self._fancy_replace_non_recursive(seq1, i1, i2, seq2, j1, j2,modified_lines,added_lines,deleted_lines,lines1,lines2)
                
            elif tag == 'delete':
                for i in range(i1, i2):
                    deleted_lines[i+1] = lines1[i]
            elif tag == 'insert':
                for i in range(j1, j2):
                    added_lines[i+1] = lines2[i]
        print("------------------------------------modified_lines------------------------------------")
        print(modified_lines)
        print("------------------------------------added_lines------------------------------------")
        print(added_lines)
        print("------------------------------------deleted_lines------------------------------------")
        print(deleted_lines)
        print("------------------------------------compare-end-------------------------------------")
        return modified_lines,added_lines,deleted_lines
    
    def compare_excel(self):
        # 初始化修改的表格的字典
        diff_infos = {}
        # 遍历共同的表格
        for sheet_name in self.common_sheets:
            modified_lines,added_lines,deleted_lines = self.diff_excel_sheet(sheet_name)
            if modified_lines:
                if sheet_name not in diff_infos:
                    diff_infos[sheet_name] = {}
                diff_infos[sheet_name]['replace'] = modified_lines
            if added_lines:
                if sheet_name not in diff_infos:
                    diff_infos[sheet_name] = {}
                diff_infos[sheet_name]['insert'] = added_lines
            if deleted_lines:
                if sheet_name not in diff_infos:
                    diff_infos[sheet_name] = {}
                diff_infos[sheet_name]['delete'] = list(deleted_lines.keys())

        return diff_infos